import pandas as pd
import string
import datetime as dt

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, NamedStyle
from openpyxl.chart import BarChart, Reference
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE

def automate_excel():
    # Get the current date
    today = dt.datetime.now()

    # Get the name of the current month
    current_month = today.strftime("%B").lower()

    # Get file name
    file_name = f"./source/sales_{current_month}.xlsx"

    # Read excel file
    excel_file = pd.read_excel(file_name)

    # Make pivot table
    pivot_table = excel_file.pivot_table(index="Gender", columns="Product line", values="Total", aggfunc="sum").round(0)
    
    # Send the report table to excel file (A4)
    pivot_table.to_excel(f"./output/report_{current_month}.xlsx", sheet_name="Report", startrow=4)

    # Loading workbook and selecting sheet
    wb = load_workbook(f"./output/report_{current_month}.xlsx")
    sheet = wb["Report"]

    # Cell references (original spreadsheet)
    min_column = wb.active.min_column
    max_column = wb.active.max_column
    min_row = wb.active.min_row
    max_row = wb.active.max_row

    # Add title and subtitle (A1 & A2)
    sheet["A1"] = "Sales Report"
    sheet["A2"] = f"{current_month}"
    sheet["A1"].font = Font("Arial", bold=True, size=20)
    sheet["A2"].font = Font("Arial", bold=True, size=10)

    # Applying formulas
    alphabet = list(string.ascii_uppercase)
    excel_alphabet = alphabet[0:max_column] # Python lists start on 0 -> A=0, B=1, C=2. #note2 the [a:b] takes b-a elements
    
    # Apply currency format to cells in column A
    currency_style = NamedStyle(name="currency_style", number_format=FORMAT_CURRENCY_USD_SIMPLE)

    # Sum in columns B-G
    for letter in excel_alphabet:
        if letter != "A":
            sheet[f"{letter}{max_row+1}"] = f"=SUM({letter}{min_row+1}:{letter}{max_row})"
            sheet[f"{letter}{max_row+1}"].style = currency_style
    sheet[f"{excel_alphabet[0]}{max_row+1}"] = "Total"

    # Adding a chart
    barchart = BarChart()

    # Define chart data and categories
    data = Reference(sheet, min_col=min_column+1, max_col=max_column, min_row=min_row, max_row=max_row) #including headers
    categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row+1, max_row=max_row) #not including headers

    # Add chart data and categories
    barchart.add_data(data, titles_from_data=True)
    barchart.set_categories(categories)

    # Select chart locations
    sheet.add_chart(barchart, "A12")
    barchart.title = "Sales by Product line"
    barchart.style = 5 # Choose the chart style

    wb.save(f"./output/report_{current_month}.xlsx")

if __name__ == "__main__":
    automate_excel()